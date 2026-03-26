# --------------------- IMPORTS ---------------------
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, WebDriverException,
    StaleElementReferenceException, NoSuchElementException
)
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from io import BytesIO
import time
import hashlib
import os
import json
import shutil
import tempfile
import uuid
import subprocess
import logging

import os
import uuid
import shutil
import tempfile
import logging
import platform

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.common.exceptions import WebDriverException


class WebDriverManager:
    def __init__(self):
        self.driver = None
        self.temp_dir = None

    def start_driver(self, headless=True):

        # ✅ 1. Réutilisation session
        if self.driver:
            try:
                _ = self.driver.title
                logging.info("⚡ Session existante OK")
                return self.driver
            except:
                logging.warning("⚠️ Session morte → restart")
                self.stop_driver()

        options = Options()

        # ✅ 2. Profil isolé
        unique_id = str(uuid.uuid4())
        self.temp_dir = os.path.join(
            tempfile.gettempdir(),
            f"chrome_{os.getpid()}_{unique_id}"
        )
        options.add_argument(f"--user-data-dir={self.temp_dir}")

        # ✅ 3. Options stabilité (ULTRA IMPORTANT)
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-extensions")
        options.add_argument("--disable-notifications")
        options.add_argument("--disable-infobars")
        options.add_argument("--disable-blink-features=AutomationControlled")

        options.add_argument("--no-first-run")
        options.add_argument("--no-default-browser-check")

        if headless:
            options.add_argument("--headless=new")
            options.add_argument("--window-size=1920,1080")

        # ✅ 4. Détection environnement
        is_linux = platform.system() == "Linux"

        try:
            logging.info("🚀 Démarrage Chrome...")

            if is_linux:
                # 🔥 STREAMLIT CLOUD

                chrome_paths = [
                    "/usr/bin/chromium",
                    "/usr/bin/chromium-browser",
                    "/usr/bin/google-chrome",
                ]

                for path in chrome_paths:
                    if os.path.exists(path):
                        options.binary_location = path
                        logging.info(f"✅ Chrome trouvé: {path}")
                        break

                chromedriver_paths = [
                    "/usr/bin/chromedriver",
                    "/usr/lib/chromium/chromedriver",
                ]

                service = None
                for path in chromedriver_paths:
                    if os.path.exists(path):
                        service = ChromeService(executable_path=path)
                        logging.info(f"✅ Driver trouvé: {path}")
                        break

                # 🔥 fallback automatique
                if service:
                    self.driver = webdriver.Chrome(service=service, options=options)
                else:
                    logging.warning("⚠️ Driver non trouvé → fallback Selenium")
                    self.driver = webdriver.Chrome(options=options)

            else:
                # 💻 LOCAL (Windows → Edge possible)
                logging.info("💻 Mode local détecté")

                try:
                    # Edge si dispo
                    from selenium.webdriver.edge.service import Service as EdgeService
                    from selenium.webdriver.edge.options import Options as EdgeOptions

                    edge_options = EdgeOptions()
                    edge_options.use_chromium = True

                    if headless:
                        edge_options.add_argument("--headless=new")

                    self.driver = webdriver.Edge(options=edge_options)
                    logging.info("✅ Edge lancé")

                except Exception as e:
                    logging.warning(f"⚠️ Edge KO → fallback Chrome ({e})")
                    self.driver = webdriver.Chrome(options=options)

            # ✅ 5. Paramètres driver
            self.driver.implicitly_wait(10)
            self.driver.set_page_load_timeout(60)

            logging.info("✅ Navigateur prêt")
            return self.driver

        except Exception as e:
            logging.error(f"❌ Erreur driver: {e}")
            self.cleanup_temp()
            return None

    def stop_driver(self):
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
            finally:
                self.driver = None
                self.cleanup_temp()
                logging.info("🛑 Driver stoppé")

    def cleanup_temp(self):
        if self.temp_dir and os.path.exists(self.temp_dir):
            try:
                shutil.rmtree(self.temp_dir, ignore_errors=True)
            except:
                pass
            self.temp_dir = None



# --- gestion_identifiants.py ---
def identifiants_cgaweb_rpe():
    # On essaie de voir si on est dans un contexte Streamlit
    try:
        import streamlit as st
        user = st.session_state.get("rpe_user")
        pwd = st.session_state.get("rpe_pass")
        secret = st.session_state.get("rpe_secret")
        if user and pwd and secret:
            return user, pwd, secret
    except: 
        pass # Si échec, on passe au fallback .env (pour le Worker)

    from dotenv import load_dotenv
    load_dotenv()
    return os.getenv("CGAWEB_USER_RPE"), os.getenv("CGAWEB_PASS_RPE"), os.getenv("CGAWEB_SECRET_RPE")



    

# --- tester_connexion.py ---
def tester_connexion_cgaweb(url, utilisateur, mot_de_passe, secret, headless=True):
    """
    Teste la connexion CGAWEB en utilisant la classe WebDriverManager.
    """

    manager = WebDriverManager()

    try:
        print(" Lancement du navigateur via WebDriverManager...")

        # Appel direct de la fonction de connexion
        driver, original_window, new_window = connexion_cgaweb(
            manager=manager,
            url=url,
            utilisateur=utilisateur,
            mot_de_passe=mot_de_passe,
            secret=secret
        )

        if driver:
            print(" Test connexion CGAWEB : SUCCÈS")
            return True
        else:
            print(" Échec de l'authentification.")
            return False

    except Exception as e:
        print(f" Erreur critique lors du test : {e}")
        return False

    finally:
        # fermeture propre du navigateur
        try:
            if manager.driver:
                try:
                    deconnexion_cgaweb(manager.driver)
                except Exception:
                    pass

                manager.stop_driver()

        except Exception:
            pass

        print("🧹 Navigateur fermé et ressources nettoyées.")

            

# génération du code otp
import pyotp

INTERVAL = 30
current_otp = None  # variable globale

def generate_otp(secret, interval=30):
    """
    Génère un OTP TOTP valide basé sur le secret fourni.
    Compatible Google Authenticator / CGAWEB.
    Met aussi à jour la variable globale current_otp.
    """
    global current_otp
    totp = pyotp.TOTP(secret, interval=interval)
    current_otp = totp.now()
    return current_otp



from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import time
import logging

def connexion_cgaweb(manager, url, utilisateur, mot_de_passe, secret=None, headless=True):
    """
    Connexion CGAWEB utilisant un WebDriverManager indépendant.
    Chaque appel crée ou réutilise une session valide, isolée.
    """
    logging.info(f"Début de connexion à CGAWEB pour {utilisateur}...")

    # --- 1. Démarre le driver via le manager ---
    driver = manager.start_driver(headless=headless)
    if not driver:
        logging.error("❌ Impossible de démarrer le driver Edge.")
        return None, None, None

    try:
        # --- 2. Navigation vers la page de login ---
        driver.get(url)

        # --- 3. Remplissage du formulaire ---
        logging.info("Remplissage du formulaire de login...")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "cuser"))).send_keys(utilisateur)
        driver.find_element(By.ID, "pass").send_keys(mot_de_passe)
        driver.find_element(By.NAME, "login").click()

        # --- 4. Vérification erreur login ---
        time.sleep(1.5)
        message_erreur = driver.find_elements(By.XPATH, "//div[contains(@style, 'color: #CE0000')]")
        if message_erreur and message_erreur[0].is_displayed():
            logging.error(f"Connexion refusée : {message_erreur[0].text.strip()}")
            return None, None, None

        # --- 5. Gestion OTP (si requis) ---
        try:
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "factors-div")))
            logging.info("Étape OTP détectée...")
            
            facteur_radio = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//tr[td//b[contains(text(),'Google Authenticator')]]//input[@type='radio']")
                )
            )
            facteur_radio.click()
            driver.find_element(By.NAME, "sendChallenge").click()

            otp_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "passCode")))

            if not secret:
                logging.error("Secret OTP non fourni.")
                return None, None, None

            otp_code = generate_otp(secret, 30)
            otp_input.send_keys(otp_code)
            driver.find_element(By.NAME, "verifyChallenge").click()
            logging.info(f"Code OTP {otp_code} injecté.")

        except TimeoutException:
            logging.info("ℹ Étape OTP ignorée (session existante ou non requise).")

        # --- 6. Navigation vers la popup 'ss' ---
        time.sleep(2)
        original_window = driver.current_window_handle
        all_windows_before = driver.window_handles

        WebDriverWait(driver, 15).until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "titleFrame")))
        driver.find_element(By.ID, "ss").click()

        # Attente ouverture nouvelle fenêtre
        WebDriverWait(driver, 15).until(EC.number_of_windows_to_be(len(all_windows_before) + 1))
        new_window = [h for h in driver.window_handles if h != original_window][0]

        logging.info("✅ Connexion réussie et nouvelle fenêtre prête.")
        return driver, original_window, new_window

    except Exception as e:
        logging.error(f"❌ Erreur technique CGAWEB : {e}")
        return None, None, None



# --- Déconnexion CGAWEB ---
def deconnexion_cgaweb(driver):
    try:
        if driver:
            print("🔌 Déconnexion de CGAWEB...")

            # --- Revenir sur la fenêtre principale ---
            if len(driver.window_handles) > 1:
                # Fermer les fenêtres secondaires
                for handle in driver.window_handles[1:]:
                    driver.switch_to.window(handle)
                    driver.close()
                driver.switch_to.window(driver.window_handles[0])

            # --- Revenir au frame principal ---
            driver.switch_to.default_content()
            driver.switch_to.frame("titleFrame")

            # --- Cliquer sur le bouton déconnexion ---
            driver.find_element(By.XPATH, "//a[contains(@href, 'DeconnexionServlet')]").click()
            print("✅ Déconnexion effectuée.")
    except Exception as e:
        print(f"❌ Problème lors de la déconnexion : {e}")
   

# --- Recherche par décodeur ---
def rechercher_par_decodeur(driver, num_decodeur, original_window, new_window):
    try:
        driver.switch_to.window(new_window)
        print("🧭 Navigué vers la fenêtre popup de recherche.")

        champ_numdec = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, "numdec"))
        )
        champ_numdec.clear()
        champ_numdec.send_keys(num_decodeur)
        print(f"🔢 Numéro de décodeur {num_decodeur} saisi.")

        driver.find_element(By.NAME, "search").click()
        print("🔍 Recherche lancée.")

        try:
            WebDriverWait(driver, 3).until(EC.alert_is_present())
            alert = driver.switch_to.alert
            message = alert.text
            alert.accept()
            print(f"⚠️ Alerte détectée : {message}")
            return None
        except TimeoutException:
            pass

        lien = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#subscriberDTABLE a"))
        )
        numero_abonne = lien.text.strip()
        lien.click()
        print(f"🔗 Lien vers l'abonné {numero_abonne} cliqué.")

        # Retour à la fenêtre principale
        driver.switch_to.window(original_window)
        driver.switch_to.default_content()
        WebDriverWait(driver, 10).until(
            EC.frame_to_be_available_and_switch_to_it((By.NAME, "_right"))
        )

        fiche_abonnes = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "view-contract"))
        )
        print("📄 Fiche d'abonné chargée avec succès.")
        return numero_abonne, fiche_abonnes

    except Exception as e:
        print(f"❌ Erreur lors de la recherche par décodeur : {e}")
        if driver:
            try:
                deconnexion_cgaweb(driver)
            except Exception as ex:
                print(f"⚠️ Problème lors de la tentative de fermeture du navigateur : {ex}")
        st.error("❌ Erreur lors de la recherche par décodeur !")
        return None


from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def extraire_abonne_et_telephones(driver):

    nom_abonne = "N/A"
    date_debut = "N/A"
    date_fin = "N/A"
    telephone_1 = "N/A"
    telephone_2 = "N/A"

    wait = WebDriverWait(driver, 20)

    try:

        # ==============================
        # 1️⃣ NOM + PERIODE (titleFrame)
        # ==============================
        driver.switch_to.default_content()

        wait.until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "titleFrame")))

        nom_elem = wait.until(EC.presence_of_element_located((By.ID, "names")))
        periode_elem = wait.until(EC.presence_of_element_located((By.ID, "period")))

        nom_abonne = " ".join(nom_elem.find_element(By.TAG_NAME, "b").text.split())

        periode_txt = periode_elem.text.replace("Période :", "").strip()
        if "-" in periode_txt:
            date_debut, date_fin = [d.strip() for d in periode_txt.split("-", 1)]

        driver.switch_to.default_content()

        # ==============================
        # 2️⃣ FRAME MENU (_right)
        # ==============================
        wait.until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "_right")))

        lien_xpath = "//a[contains(@href,'modaddress')]"

        lien_adresse = wait.until(
            EC.presence_of_element_located((By.XPATH, lien_xpath))
        )

        # scroll + click JS (plus fiable en headless)
        driver.execute_script("arguments[0].scrollIntoView(true);", lien_adresse)
        driver.execute_script("arguments[0].click();", lien_adresse)

        # ==============================
        # 3️⃣ ATTENTE FORMULAIRE
        # ==============================
        wait.until(EC.presence_of_element_located((By.NAME, "address")))

        def reconstruire_numero(name):
            elements = driver.find_elements(By.XPATH, f"//input[@name='{name}']")
            valeurs = []

            for el in elements:
                val = el.get_attribute("value")
                if val and val.strip():
                    valeurs.append(val.strip())

            return "".join(valeurs) if valeurs else "N/A"

        telephone_1 = reconstruire_numero("mobile1")
        telephone_2 = reconstruire_numero("mobile2")

        driver.switch_to.default_content()

    except Exception as e:
        print("❌ Erreur extraction abonné et téléphones :", e)
        driver.switch_to.default_content()

    return nom_abonne, date_debut, date_fin, telephone_1, telephone_2


# --- ----------------Extraction infos abonné ---

from datetime import datetime

def extraire_infos_abonne(driver, num_decodeur, numero_abonne):
    wait = WebDriverWait(driver, 20)

    infos = {
        "numero_decodeur": num_decodeur,
        "nom_abonne": "Inconnu",
        "numero_abonne": numero_abonne,
        "date_recrutement": "",
        "num_distributeur": "",
        "nom_distributeur": "",
        "telephone": "",
        "telephone_2": "",
        "fin_abonnement": "",
        "statut": "Inconnu",
        "detail-statut": "",
        "temps_traitement (sec)": "",
        "temps_moyen_traitement (sec)": "",
    }

    try:
        # ===============================
        # 1️⃣ Se positionner sur la fiche
        # ===============================
        driver.switch_to.default_content()
        wait.until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "_right")))

        # ===============================
        # 2️⃣ Infos distributeur
        # ===============================
        try:
            dist_id = wait.until(EC.presence_of_element_located((By.ID, "numdist")))
            infos["num_distributeur"] = dist_id.text.strip()
        except:
            pass

        try:
            dist_nom = driver.find_element(By.ID, "nomdist")
            infos["nom_distributeur"] = dist_nom.text.strip()
        except:
            pass

        # ===============================
        # 3️⃣ Nom, téléphones et dates via ta fonction
        # ===============================
        try:
            nom_abonne, date_debut, date_fin, tel1, tel2 = extraire_abonne_et_telephones(driver)
            infos["nom_abonne"] = nom_abonne if nom_abonne else "Inconnu"
            infos["date_recrutement"] = date_debut if date_debut else ""
            infos["fin_abonnement"] = date_fin if date_fin else ""
            infos["telephone"] = tel1 if tel1 else ""
            infos["telephone_2"] = tel2 if tel2 else ""
        except Exception as e:
            print(f" Erreur extraction abonné/telephones via extraire_abonne_et_telephones : {e}")

        # ===============================
        # 4️⃣ Calcul statut
        # ===============================
        if infos["fin_abonnement"] and "/" in infos["fin_abonnement"]:
            try:
                date_fin_dt = datetime.strptime(infos["fin_abonnement"], "%d/%m/%Y").date()
                today = datetime.today().date()
                delta = (date_fin_dt - today).days

                if delta > 0:
                    infos["statut"] = "Actif"
                    infos["detail-statut"] = f"Reste {delta} jours"
                elif delta == 0:
                    infos["statut"] = "Actif"
                    infos["detail-statut"] = "Expire aujourd'hui"
                else:
                    infos["statut"] = "Échu"
                    infos["detail-statut"] = f"Expiré depuis {abs(delta)} jours"

            except:
                infos["statut"] = "Erreur Date"
        else:
            infos["detail-statut"] = "Pas de date de fin"

        return infos

    except Exception as e:
        print(f" Erreur critique extraction : {e}")
        return infos



from datetime import datetime

def extraire_infos_abonne_rapide(fiche_abonne, num_decodeur, driver, numero_abonne, champs_choisis):
    """
    Extrait de manière sélective les informations demandées.
    """
    infos = {"numero_decodeur": str(num_decodeur), "numero_abonne": str(numero_abonne)}
    
    # Liste des champs gérés par la fonction extraire_telephone
    champs_externes = ["nom_abonne", "date_recrutement", "fin_abonnement", "telephone", "telephone_2"]
    besoin_extraire_tel = any(f in champs_choisis for f in champs_externes)

    try:
        # --- 1. Infos présentes directement sur la fiche (Distributeur) ---
        if "num_distributeur" in champs_choisis:
            try:
                infos["num_distributeur"] = fiche_abonne.find_element(By.ID, "numdist").text.strip()
            except:
                infos["num_distributeur"] = "N/A"

        if "nom_distributeur" in champs_choisis:
            try:
                infos["nom_distributeur"] = fiche_abonne.find_element(By.ID, "nomdist").text.strip()
            except:
                infos["nom_distributeur"] = "N/A"

        # --- 2. Infos nécessitant de changer de frame (via extraire_telephone) ---
        if besoin_extraire_tel:
            try:
                # On récupère tout d'un coup pour limiter les changements de frames
                tel, nom, debut, fin, tel2 = extraire_abonne_et_telephones(driver, fiche_abonne)
                
                mapping = {
                    "telephone": tel,
                    "nom_abonne": nom,
                    "date_recrutement": debut,
                    "fin_abonnement": fin,
                    "telephone_2": tel2
                }
                
                for champ in champs_externes:
                    if champ in champs_choisis:
                        infos[champ] = mapping.get(champ, "N/A")
            except Exception as e:
                print(f"⚠️ Erreur frame/telephone pour {num_decodeur}: {e}")
                for champ in champs_externes:
                    if champ in champs_choisis: infos[champ] = "Erreur"

        # --- 3. Calcul dynamique du statut ---
        if "statut" in champs_choisis or "detail_statut" in champs_choisis:
            fin_abonne = infos.get("fin_abonnement")
            if fin_abonne and "/" in str(fin_abonne):
                try:
                    date_fin_dt = datetime.strptime(fin_abonne, "%d/%m/%Y").date()
                    today = datetime.now().date()
                    delta = (date_fin_dt - today).days

                    if delta > 0:
                        infos["statut"] = "Actif"
                        infos["detail_statut"] = f"Reste {delta} jours"
                    elif delta == 0:
                        infos["statut"] = "Actif"
                        infos["detail_statut"] = "Expire aujourd'hui"
                    else:
                        infos["statut"] = "Échu"
                        infos["detail_statut"] = f"Expiré depuis {abs(delta)} jours"
                except:
                    infos["statut"] = "Inconnu"
                    infos["detail_statut"] = "Format date invalide"
            else:
                infos["statut"] = "Inconnu"
                infos["detail_statut"] = "Date absente"

        return infos

    except Exception as e:
        print(f"❌ Erreur globale extraction {num_decodeur}: {e}")
        return infos


import time
import streamlit as st

def traiter_decodeurs_rapide(driver, df, original_window, new_window, champs_choisis, statut_txt=None, progress_bar=None):
    """
    Parcourt la liste des décodeurs et extrait les informations.
    Optimisé pour la vitesse et la stabilité du driver.
    """
    resultats = []
    traitements = []
    # Nettoyage de la liste des numéros (suppression des doublons et vides)
    liste_decodeurs = df["numéro de décodeur"].dropna().unique().tolist()
    total = len(liste_decodeurs)

    for i, num in enumerate(liste_decodeurs, 1):
        # 1. Vérification de l'arrêt d'urgence
        if st.session_state.get("stop_processing", False):
            if statut_txt:
                statut_txt.warning("⏸️ Traitement interrompu par l'utilisateur.")
            break

        debut = time.time()
        num_str = str(num).strip()
        
        try:
            if statut_txt:
                statut_txt.info(f"🔍 Recherche décodeur {i}/{total} : {num_str}")

            # 2. Appel à la recherche
            res = rechercher_par_numero(driver, num_str, original_window, new_window)
            
            if res:
                numero_abonne, fiche_abonne = res
                # 3. Appel à l'extraction rapide (version corrigée précédemment)
                infos = extraire_infos_abonne_rapide(fiche_abonne, num_str, driver, numero_abonne, champs_choisis)
                
                if infos:
                    tps = round(time.time() - debut, 2)
                    traitements.append(tps)
                    infos["temps_traitement (sec)"] = tps
                    resultats.append(infos)
                    print(f"✅ [{i}/{total}] Succès : {num_str}")
            else:
                # 4. Gestion du cas "Non trouvé"
                tps = round(time.time() - debut, 2)
                traitements.append(tps)
                
                # On initialise avec "Non trouvé" pour tous les champs demandés
                infos_echec = {f: "Non trouvé" for f in champs_choisis}
                infos_echec.update({
                    "numero_decodeur": num_str,
                    "numero_abonne": "Non trouvé",
                    "temps_traitement (sec)": tps
                })
                resultats.append(infos_echec)
                print(f"⚠️ [{i}/{total}] Décodeur introuvable : {num_str}")

        except Exception as e:
            print(f"❌ Erreur critique sur décodeur {num_str} : {e}")
            # On ajoute quand même une ligne d'erreur pour ne pas décaler les résultats
            resultats.append({"numero_decodeur": num_str, "statut": "Erreur technique"})

        # 5. Mise à jour de la barre de progression Streamlit
        if progress_bar:
            progress_bar.progress(i / total)

    # 6. Calcul du temps moyen final
    if traitements:
        temps_moyen = round(sum(traitements) / len(traitements), 2)
        for r in resultats:
            r["temps_moyen_traitement (sec)"] = temps_moyen
        
        if statut_txt:
            statut_txt.success(f"✨ Traitement terminé. Temps moyen : {temps_moyen}s / décodeur.")

    return resultats

from openpyxl.styles import PatternFill
from io import BytesIO
from datetime import datetime
import pandas as pd



def resultatsmart(resultats, champs_choisis):
    """
    Génère le fichier Excel avec uniquement les champs extraits et applique un code couleur.
    """
    now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    # On nettoie le nom de fichier pour éviter les erreurs système
    nom_fichier = f"Extraction_CGAWEB_{now}.xlsx"

    # On construit la liste des colonnes en évitant les doublons
    colonnes_fixes_debut = ["numero_decodeur", "numero_abonne"]
    colonnes_fixes_fin = ["temps_traitement (sec)", "temps_moyen_traitement (sec)"]
    
    # On filtre champs_choisis pour ne pas répéter les colonnes déjà présentes dans fixes_debut
    colonnes_milieu = [c for c in champs_choisis if c not in colonnes_fixes_debut]
    
    colonnes_finales = colonnes_fixes_debut + colonnes_milieu + colonnes_fixes_fin

    # Uniformiser chaque résultat pour éviter les KeyError
    donnees_uniformes = []
    for r in resultats:
        # On utilise r.get(col, "") pour remplir par du vide si la donnée manque
        ligne = {col: r.get(col, "") for col in colonnes_finales}
        donnees_uniformes.append(ligne)

    df = pd.DataFrame(donnees_uniformes, columns=colonnes_finales)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Résultats')
        ws = writer.sheets['Résultats']

        # Définition des styles de remplissage
        vert = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Vert clair
        rouge = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Rouge clair
        jaune = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Jaune clair

        # Application du formatage conditionnel manuel
        if "statut" in df.columns:
            col_statut_idx = df.columns.get_loc("statut") + 1
            col_finab_idx = df.columns.get_loc("fin_abonnement") + 1 if "fin_abonnement" in df.columns else None

            today = datetime.now().date()

            for row_idx in range(2, len(df) + 2):
                statut_cell = ws.cell(row=row_idx, column=col_statut_idx)
                valeur_statut = str(statut_cell.value)

                # Logique de couleur
                if valeur_statut == "Actif":
                    # Vérification si l'échéance est proche (<= 5 jours)
                    est_proche = False
                    if col_finab_idx:
                        val_date = ws.cell(row=row_idx, column=col_finab_idx).value
                        try:
                            # Gestion flexible de la date (string ou datetime)
                            if isinstance(val_date, str):
                                date_fin = datetime.strptime(val_date, "%d/%m/%Y").date()
                            else:
                                date_fin = val_date.date()
                            
                            if (date_fin - today).days <= 5:
                                est_proche = True
                        except:
                            pass
                    
                    statut_cell.fill = jaune if est_proche else vert
                
                elif valeur_statut == "Échu":
                    statut_cell.fill = rouge

        # Ajustement automatique de la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column_letter].width = max_length + 2

    output.seek(0)
    return output, nom_fichier




def traiter_decodeurs(driver, df, colonne_source, original_window, new_window, statut_txt=None, progress_bar=None):

    resultats = []
    traitements = []
    source = df[colonne_source].dropna().astype(str).str.strip()
    total = len(source)

    if total == 0:
        st.error("La colonne sélectionnée est vide.")
        return []

    st.session_state.stop_processing = False

    for i, valeur in enumerate(source, 1):

        if st.session_state.get("stop_processing", False):
            break

        debut = time.time()
        valeur_str = valeur.strip()

        try:
            # 🔁 Toujours repartir proprement
            driver.switch_to.window(new_window)
            driver.switch_to.default_content()

            # ===============================
            # 1️⃣ Recherche
            # ===============================
            numero_trouve = rechercher_par_numero(
                driver,
                valeur_str,
                original_window,
                new_window
            )

            if numero_trouve:

                if colonne_source.lower() == "abonne":
                    numero_abonne = valeur_str
                    numero_decodeur = numero_trouve
                else:
                    numero_decodeur = valeur_str
                    numero_abonne = numero_trouve

                # ===============================
                # 2️⃣ Extraction directe (sans fiche_abonne)
                # ===============================
                infos = extraire_infos_abonne(
                    driver,
                    numero_decodeur,
                    numero_abonne
                )

                if infos:
                    tps = round(time.time() - debut, 2)
                    traitements.append(tps)
                    infos["temps_traitement (sec)"] = tps
                    resultats.append(infos)

                    st.session_state.resultats_en_cours = resultats
                    maj_progression(i, total, statut_txt, progress_bar)
                    continue

            # ===============================
            # Cas Non trouvé
            # ===============================
            tps = round(time.time() - debut, 2)
            traitements.append(tps)

            placeholder = {
                "numero_decodeur": valeur_str if colonne_source.lower() != "abonne" else "Non trouvé",
                "numero_abonne": valeur_str if colonne_source.lower() == "abonne" else "Non trouvé",
                "nom_abonne": "Introuvable",
                "statut": "Échu",
                "detail-statut": "Aucune fiche trouvée",
                "temps_traitement (sec)": tps
            }

            resultats.append(placeholder)
            st.session_state.resultats_en_cours = resultats
            maj_progression(i, total, statut_txt, progress_bar)

        except Exception as e:
            print(f"⚠️ Erreur sur {valeur_str} : {e}")
            tps = round(time.time() - debut, 2)
            resultats.append({
                "numero_decodeur": valeur_str,
                "numero_abonne":   "Erreur",
                "nom_abonne":      "Erreur technique",
                "statut":          "Erreur",
                "detail-statut":   str(e),
                "temps_traitement (sec)": tps,
            })
            st.session_state.resultats_en_cours = resultats 
            maj_progression(i, total, statut_txt, progress_bar)

    return resultats



# fonction de progression
def maj_progression(i, total, statut_txt=None, progress_bar=None, sleep_time=0.05):
    """
    Met à jour le texte de statut et la barre de progression.

    i : numéro d'élément en cours de traitement (1-based)
    total : nombre total d'éléments à traiter
    statut_txt : élément Streamlit text() (optionnel)
    progress_bar : élément Streamlit progress() (optionnel)
    sleep_time : temps d'attente pour forcer le rafraîchissement UI
    """
    if statut_txt:
        pourcentage = int((i / total) * 100)
        statut_txt.text(f"Traitement : {pourcentage}% ({i} sur {total})")

    if progress_bar:
        progress_bar.progress(i / total)
        time.sleep(sleep_time)



from concurrent.futures import ThreadPoolExecutor
import streamlit as st

def traiter_decodeurs_multi_navigateurs(df, url, utilisateur, mot_de_passe, champs_choisis, statut_txt=None, progress_bar=None, session_state=None):
    """
    Traitement multi-navigateurs optimisé.
    Chaque navigateur traite une partie du DataFrame.
    """
    # ⚙️ Priorité aux identifiants RPE
    if session_state:
        u_rpe = session_state.get("rpe_user") # Harmonisé avec votre code précédent
        p_rpe = session_state.get("rpe_pass")
        if u_rpe and p_rpe:
            utilisateur, mot_de_passe = u_rpe, p_rpe

    # Séparation intelligente des données
    total_lignes = len(df)
    mid = total_lignes // 2
    df_part1 = df.iloc[:mid].reset_index(drop=True)
    df_part2 = df.iloc[mid:].reset_index(drop=True)

    resultats_globaux = []

    # Initialisation résultats partiels sécurisée
    if session_state is not None and "resultats_en_cours" not in session_state:
        session_state.resultats_en_cours = []

    # Fonction de travail pour un thread
    def executer_navigation(sous_df, thread_id):
        res_thread = []
        driver = None
        try:
            print(f"🚀 Thread {thread_id} : Lancement du navigateur...")
            driver, orig_win, new_win = connexion_cgaweb(url, utilisateur, mot_de_passe, headless=True)
            
            if driver:
                # On appelle la version RAPIDE corrigée précédemment
                res_thread = traiter_decodeurs_rapide(
                    driver, sous_df, orig_win, new_win, champs_choisis, 
                    statut_txt=None, progress_bar=None # Pas de widgets ST dans les threads !
                )
                
                if session_state is not None:
                    session_state.resultats_en_cours.extend(res_thread)
            
            return res_thread
        except Exception as e:
            print(f"❌ Erreur Thread {thread_id} : {e}")
            return []
        finally:
            if driver:
                try: deconnexion_cgaweb(driver)
                except: pass

    # --- Lancement du ThreadPool ---
    # Note : Streamlit ne peut pas mettre à jour progress_bar depuis ici
    if statut_txt:
        statut_txt.info("⚡ Traitement parallèle lancé sur 2 navigateurs...")

    with ThreadPoolExecutor(max_workers=2) as executor:
        # On lance les deux tâches
        future1 = executor.submit(executer_navigation, df_part1, 1)
        future2 = executor.submit(executer_navigation, df_part2, 2)

        # On attend les résultats
        res1 = future1.result()
        res2 = future2.result()

    resultats_globaux = res1 + res2

    # Mise à jour finale de l'interface
    if progress_bar:
        progress_bar.progress(1.0)
    if statut_txt:
        statut_txt.success(f"✅ Multi-traitement terminé : {len(resultats_globaux)} décodeurs traités.")

    return resultats_globaux


# --- Export Excel ---
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl.styles import PatternFill

def resultat(resultats):
    """
    Génère un fichier Excel complet avec formatage conditionnel des statuts.
    Version optimisée pour l'export final.
    """
    if not resultats:
        return None, "aucun_resultat.xlsx"

    now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    nom_fichier = f"resultats_cgaweb_{now}.xlsx"
    
    # Création du DataFrame
    df = pd.DataFrame(resultats)

    # Définition de l'ordre strict des colonnes pour l'export final
    colonnes_finales = [
        "numero_decodeur", "nom_abonne", "numero_abonne", "date_recrutement",
        "num_distributeur", "nom_distributeur", "telephone", "telephone_2",
        "fin_abonnement", "statut", "detail-statut",
        "temps_traitement (sec)", "temps_moyen_traitement (sec)",
    ]

    # Ajout des colonnes manquantes et réordonnancement
    for col in colonnes_finales:
        if col not in df.columns:
            df[col] = "N/A"
    
    df = df[colonnes_finales]

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Résultats')
        ws = writer.sheets['Résultats']

        # --- Styles de cellules ---
        vert = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # Actif OK
        rouge = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Échu
        jaune = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Actif critique (<= 5j)

        # Identification des indices de colonnes (1-based pour openpyxl)
        col_statut_idx = colonnes_finales.index("statut") + 1
        col_finab_idx = colonnes_finales.index("fin_abonnement") + 1
        
        today = datetime.now().date()

        # --- Parcours des lignes pour le formatage ---
        for row_idx in range(2, len(df) + 2):
            statut_cell = ws.cell(row=row_idx, column=col_statut_idx)
            finab_cell = ws.cell(row=row_idx, column=col_finab_idx)
            
            jours_restant = None
            
            # 1. Calcul des jours restants de manière sécurisée
            if finab_cell.value and str(finab_cell.value).strip() not in ["", "N/A"]:
                try:
                    # Conversion flexible via pandas
                    date_fin = pd.to_datetime(finab_cell.value, dayfirst=True).date()
                    jours_restant = (date_fin - today).days
                except Exception:
                    jours_restant = None

            # 2. Application des couleurs selon le statut et l'urgence
            statut_val = str(statut_cell.value).strip()
            
            if statut_val == "Actif":
                if jours_restant is not None and jours_restant <= 5:
                    statut_cell.fill = jaune
                else:
                    statut_cell.fill = vert
            elif statut_val == "Échu":
                statut_cell.fill = rouge

        # --- Ajustement automatique de la largeur des colonnes ---
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column_letter].width = max_length + 2

    output.seek(0)
    return output, nom_fichier




import pandas as pd
import os
import time

def sauvegarder_resultats_intermediaires(session_state, chemin="resultats_partiels.xlsx"):
    """
    Sauvegarde les résultats partiels en cours dans un fichier Excel.
    Utilisé pour l'arrêt d'urgence ou pour la sauvegarde périodique.
    """
    if not session_state:
        return None

    # Vérifie qu'il y a des résultats à sauvegarder
    resultats = session_state.get("resultats_en_cours", [])
    if not resultats:
        print("⚠️ Aucun résultat partiel à sauvegarder.")
        return None

    # Convertir en DataFrame
    df_partiel = pd.DataFrame(resultats)

    # Ajouter timestamp si le fichier existe déjà
    if os.path.exists(chemin):
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        chemin = chemin.replace(".xlsx", f"_{timestamp}.xlsx")

    # Sauvegarde en Excel
    try:
        df_partiel.to_excel(chemin, index=False)
        print(f"💾 Résultats partiels sauvegardés dans : {chemin}")
        return chemin
    except Exception as e:
        print(f"❌ Erreur sauvegarde résultats partiels : {e}")
        return None

    
# --------------------- ARRET D'URGENCE ---------------------
def arret_urgence(driver=None):
    """
    Arrêt d'urgence sécurisé :
    - Sauvegarde les résultats déjà extraits
    - Déconnexion CGAWEB si possible
    - Fermeture propre du navigateur
    """

    st.warning("⛔ Arrêt d'urgence déclenché. Sauvegarde en cours...")

    # 1️⃣ Récupération sécurisée des résultats en cours
    resultats_en_cours = st.session_state.get("resultats_en_cours", [])

    if resultats_en_cours:
        try:
            sauvegarder_resultats_intermediaires(resultats_en_cours)
            st.success("💾 Résultats partiels sauvegardés avec succès")
        except Exception as e:
            st.error(f"⚠ Erreur lors de la sauvegarde : {e}")
    else:
        st.info("ℹ Aucun résultat à sauvegarder")

    # 2️⃣ Déconnexion CGAWEB (sans fermer le navigateur)
    if driver:
        try:
            deconnexion_cgaweb(driver)
        except Exception:
            pass  # On ne bloque jamais l'arrêt d'urgence

    # 3️⃣ Fermeture du navigateur (UNE SEULE FOIS)
    if driver:
        try:
            driver.quit()
            st.info("🧹 Navigateur fermé correctement")
        except Exception:
            pass

    # 4️⃣ Nettoyage du state Streamlit
    for key in [
        "driver",
        "resultats_en_cours",
        "stop_requested",
        "progression"
    ]:
        if key in st.session_state:
            del st.session_state[key]

    st.success("✅ Arrêt d'urgence terminé proprement")


# --------------------- FONCTION DE RECHERCHE AVEC NUMERO D'ABONNE ---------------------
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

def rechercher_par_abonne(driver, num_abonne, original_window, new_window):
    try:
        # 1️⃣ Basculer sur la fenêtre de recherche (Popup)
        driver.switch_to.window(new_window)
        driver.switch_to.default_content()
        
        # 2️⃣ Remplir le formulaire
        print(f"🔍 Recherche de l'abonné : {num_abonne}")
        wait = WebDriverWait(driver, 10)
        
        champ_numabo = wait.until(EC.presence_of_element_located((By.NAME, "numabo")))
        champ_numabo.clear()
        champ_numabo.send_keys(num_abonne)
        
        # Clic sur le bouton Search
        btn_search = driver.find_element(By.NAME, "search")
        driver.execute_script("arguments[0].click();", btn_search)

        # 3️⃣ Gérer l'alerte potentielle (Abonné inexistant)
        try:
            WebDriverWait(driver, 3).until(EC.alert_is_present())
            alert = driver.switch_to.alert
            message = alert.text
            alert.accept()
            print(f"⚠️ Alerte CGAWEB : {message}")
            return None # On arrête ici si l'abonné n'existe pas
        except TimeoutException:
            pass 

        # 4️⃣ Attendre et cliquer sur le lien du résultat dans le tableau
        try:
            # On attend que le tableau de résultats soit chargé
            lien_resultat = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#subscriberDTABLE a")))
            numero_trouve = lien_resultat.text.strip()
            
            # Clic forcé via JS pour éviter les erreurs d'interception
            driver.execute_script("arguments[0].click();", lien_resultat)
            print(f"✅ Lien pour l'abonné {numero_trouve} cliqué.")
        except TimeoutException:
            print(f"❌ Aucun résultat trouvé dans le tableau pour {num_abonne}")
            return None

        # 5️⃣ Revenir à la fenêtre principale pour afficher la fiche
        # Sur CGAWEB, cliquer dans la popup recharge la fiche dans la fenêtre principale
        driver.switch_to.window(original_window)
        driver.switch_to.default_content()
        
        # Attendre que le frame de droite (où s'affiche la fiche) soit prêt
        wait.until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "_right")))
        
        # Localiser la fiche du contrat
        fiche_abonnes = wait.until(EC.presence_of_element_located((By.ID, "view-contract")))
        
        return numero_trouve, fiche_abonnes

    except Exception as e:
        print(f"❌ Erreur critique rechercher_par_abonne : {e}")
        return None

#-------------la fonction de recherche par décodeur ou par numero d'abonne -------------

def rechercher_par_numero(driver, numero, original_window, new_window):
    """
    Recherche générique par décodeur ou abonné.
    Version ultra-stable compatible CGAWEB (anti-stale, anti-crash).
    Retourne uniquement des valeurs simples (jamais de WebElement).
    """

    if driver is None:
        print("❌ Driver inexistant.")
        return None

    numero_str = str(numero).strip()
    type_recherche = None

    # ==============================
    # 1️⃣ Déterminer type recherche
    # ==============================
    if len(numero_str) == 14 and numero_str.startswith("00237"):
        return rechercher_par_telephone(driver, numero_str, new_window)

    elif len(numero_str) == 14:
        type_recherche = "decodeur"
        champ_name = "numdec"

    elif len(numero_str) == 8:
        type_recherche = "abonne"
        champ_name = "numabo"

    else:
        print(f"❌ Numéro invalide : {numero_str}")
        return None

    try:
        wait = WebDriverWait(driver, 20)

        # ==============================
        # 2️⃣ Aller sur popup recherche
        # ==============================
        driver.switch_to.window(new_window)
        driver.switch_to.default_content()

        champ = wait.until(
            EC.element_to_be_clickable((By.NAME, champ_name))
        )

        champ.clear()
        champ.send_keys(numero_str)

        wait.until(
            EC.element_to_be_clickable((By.NAME, "search"))
        ).click()

        # ==============================
        # 3️⃣ Gestion alerte "inexistant"
        # ==============================
        try:
            WebDriverWait(driver, 2).until(EC.alert_is_present())
            alert = driver.switch_to.alert
            print(f"⚠️ CGAWEB dit : {alert.text}")
            alert.accept()
            return None
        except TimeoutException:
            pass

        # ==============================
        # 4️⃣ Cliquer résultat
        # ==============================
        try:
            lien = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "#subscriberDTABLE a"))
            )

            numero_abonne = None
            if type_recherche == "decodeur":
                numero_abonne = lien.text.replace(" ", "").strip()

            # clic JS plus stable
            driver.execute_script("arguments[0].click();", lien)

        except TimeoutException:
            print(f"⚠️ Aucun résultat pour {numero_str}")
            return None

        # ==============================
        # 5️⃣ Retour fenêtre principale
        # ==============================
        driver.switch_to.window(original_window)
        driver.switch_to.default_content()

        # attendre que les frames soient recréés
        wait.until(lambda d: len(d.find_elements(By.TAG_NAME, "frame")) > 0)

        driver.switch_to.default_content()

        # ==============================
        # 6️⃣ Attendre frame _right
        # ==============================
        wait.until(
            EC.frame_to_be_available_and_switch_to_it((By.NAME, "_right"))
        )

        # attendre que le contrat soit visible
        wait.until(
            EC.visibility_of_element_located((By.ID, "view-contract"))
        )

        # ==============================
        # 7️⃣ Si recherche abonné → extraire décodeur
        # ==============================
        if type_recherche == "abonne":
            try:
                lien_dec = wait.until(
                    EC.presence_of_element_located(
                        (By.XPATH, "//a[@title='Origine matériel']")
                    )
                )
                numero_decodeur = lien_dec.text.replace(" ", "").strip()
                return numero_decodeur

            except TimeoutException:
                return None

        # ==============================
        # 8️⃣ Retour standard décodeur
        # ==============================
        return numero_abonne

    except Exception as e:
        print(f"❌ Erreur critique recherche {numero_str} : {e}")

        if "session id" in str(e).lower():
            raise e

        return None

    finally:
        # sécurité : revenir sur fenêtre principale
        try:
            driver.switch_to.window(original_window)
            driver.switch_to.default_content()
        except:
            pass





# --------------------- FONCTIONS DE RECHERCHE AVEC NUMERO TELEPHONE ---------------------

def rechercher_par_telephone(driver, numero_str, new_window, timeout=15):
    """
    Recherche un ou plusieurs abonnés via un numéro de téléphone sur CGAWEB.
    Récupère les données directement depuis la variable JavaScript subscriberT.
    """
    try:
        # 🔹 Basculer sur la popup
        driver.switch_to.window(new_window)
        driver.switch_to.default_content()
        print("Formulaire de recherche par téléphone trouvé.")

        # 🔹 Remplir le champ téléphone
        champ_telephone = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.NAME, "phonenumber"))
        )
        champ_telephone.clear()
        champ_telephone.send_keys(numero_str)
        print(f"Numéro de téléphone {numero_str} saisi.")

        # 🔹 Cliquer sur "search"
        driver.find_element(By.NAME, "search").click()
        print("Recherche par téléphone lancée.")

        # 🔹 Vérifier les alertes éventuelles
        try:
            WebDriverWait(driver, 5).until(EC.alert_is_present())
            alert = driver.switch_to.alert
            message = alert.text
            alert.accept()
            print(f"Alerte détectée : {message}")
            return {"erreur": message}
        except TimeoutException:
            pass

        # 🔹 Attendre que la variable JS subscriberT soit chargée
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script("return typeof subscriberT !== 'undefined' && subscriberT.length >= 0;")
        )

        # 🔹 Extraire les données depuis subscriberT
        subscriberT = driver.execute_script("return subscriberT;")
        if not subscriberT:
            print(f"Aucun abonné trouvé pour le numéro {numero_str}.")
            return []

        resultats = []
        for ligne in subscriberT:
            # ligne est une liste avec toutes les colonnes
            donnees = {
                "abonne": ligne[0],
                "c": ligne[1],
                "nom": ligne[2],
                "prenom": ligne[3],
                "some_checkbox": ligne[5],  # indwholesale
                "pays": ligne[6],
                "code_postal": ligne[7],
                "ville": ligne[8],
                "adresse": ligne[9],
                "option_majeure": ligne[10],
                "annule": ligne[12],
                "ret": ligne[11],
                "cl": ligne[13],
                "societe": ligne[14]
            }
            resultats.append(donnees)

        print(f"{len(resultats)} abonnés trouvés pour le numéro {numero_str}.")
        return resultats

    except Exception as e:
        print(f"❌ Erreur recherche par téléphone : {e}")
        return {"erreur": str(e)}

#----------------------Fonction pour le suivi abonné----------------

def extraction_suivi_abonne(fiche_abonne, driver):
    try:
        print(f"🔍 Extraction du suivi abonné pour : {fiche_abonne}")

        # 1. Cliquer sur "Suivi abonné" à partir du fragment donné
        lien_suivi = fiche_abonne.find_element(By.LINK_TEXT, "Suivi abonné")
        lien_suivi.click()

        # 2. Attendre que le fragment formContainer apparaisse
        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.ID, "formContainer"))
        )
        print("✅ Section 'Suivi abonné' chargée")

        # 3. Vérifier la pagination
        div_followup = driver.find_element(By.ID, "divFOLLOWUP")
        pages_elements = div_followup.find_elements(By.TAG_NAME, "a")
        pages = set()
        for elem in pages_elements:
            txt = elem.text.strip(">< ")
            if txt.isdigit():
                pages.add(int(txt))
        total_pages = max(pages) if pages else 1
        print(f"🔢 Nombre de pages : {total_pages}")

        all_suivis = []

        # 4. Parcourir les pages
        for page in range(1, total_pages + 1):
            print(f"📄 Traitement de la page {page}")

            if page != 1:
                try:
                    lien_page = driver.find_element(By.XPATH, f"//div[@id='divFOLLOWUP']//a[contains(text(), '{page}')]")
                    lien_page.click()
                    time.sleep(2)
                except:
                    print(f"⚠️ Impossible de naviguer vers la page {page}")
                    continue

            # Attendre le tableau principal
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "FOLLOWUPDTABLE"))
            )
            tableau = driver.find_element(By.ID, "FOLLOWUPDTABLE")
            lignes = tableau.find_elements(By.TAG_NAME, "tr")

            for ligne in lignes[1:]:  # ignorer l'en-tête
                try:
                    lien_detail = ligne.find_element(By.TAG_NAME, "a")
                    lien_detail.click()
                    time.sleep(1)

                    WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.ID, "CONTEXTDTABLE"))
                    )
                    tableau_detail = driver.find_element(By.ID, "CONTEXTDTABLE")
                    lignes_detail = tableau_detail.find_elements(By.TAG_NAME, "tr")

                    details = {}
                    for tr in lignes_detail:
                        tds = tr.find_elements(By.TAG_NAME, "td")
                        if len(tds) >= 2:
                            cle = tds[0].text.strip()
                            valeur = tds[1].text.strip()
                            details[cle] = valeur
                    all_suivis.append(details)

                except Exception as e:
                    print(f"⚠️ Erreur lors de l'extraction d'un détail : {e}")
                    continue

        print(f"✅ Extraction terminée ({len(all_suivis)} entrées)")
        return all_suivis

    except Exception as e:
        print(f"❌ Erreur dans extract_suivi_abonne : {e}")
        return []
#---------------afficher suivi abonné-----------------
def afficher_suivi_abonne(all_suivis):
    # Pagination
    nb_par_page = 10
    nb_total = len(all_suivis)
    nb_pages = (nb_total - 1) // nb_par_page + 1

    # Choisir la page
    page = st.number_input("Page", min_value=1, max_value=nb_pages, step=1, value=1)

    debut = (page - 1) * nb_par_page
    fin = min(debut + nb_par_page, nb_total)

    for i, mouvement in enumerate(all_suivis[debut:fin], start=debut + 1):
        with st.expander(f"🔸 Mouvement {i} - {mouvement.get('Date', 'Date inconnue')}"):
            for cle, valeur in mouvement.items():
                st.markdown(f"**{cle}** : {valeur}")

    # Navigation
    col1, col2 = st.columns([1, 1])
    with col1:
        if page > 1:
            st.button("⬅️ Page précédente", on_click=st.session_state.update, args=("page_suivi", page - 1))
    with col2:
        if page < nb_pages:
            st.button("➡️ Page suivante", on_click=st.session_state.update, args=("page_suivi", page + 1))

import os
import hashlib
import mysql.connector
import pandas as pd
import streamlit as st
from cryptography.fernet import Fernet


# =========================================================
# 🔌 CONNEXION
# =========================================================
import streamlit as st
import mysql.connector

conn = mysql.connector.connect(
    host=st.secrets["DB_HOST"],
    user=st.secrets["DB_USER"],
    password=st.secrets["DB_PASSWORD"],
    database=st.secrets["DB_NAME"],
    port=3306
)


# =========================================================
# 🔐 HACHAGE (mots de passe utilisateurs — irréversible)
# =========================================================
def hash_mot_de_passe(mdp):
    return hashlib.sha256(mdp.encode("utf-8")).hexdigest()


# =========================================================
# 🔒 CHIFFREMENT AES/Fernet (comptes CGAWEB — réversible)
#
# La clé doit être dans une variable d'environnement.
# Pour générer une clé une seule fois :
#   python -c "from cryptography.fernet import Fernet; print(Fernet.generate_key().decode())"
# Puis définir :
#   Windows  → set VISUV_SECRET_KEY=<valeur>
#   .env     → VISUV_SECRET_KEY=<valeur>  (avec python-dotenv)
# =========================================================
def _get_fernet() -> Fernet:
    """Retourne l'instance Fernet initialisée avec la clé secrète."""
    cle = os.environ.get("VISUV_SECRET_KEY")
    if not cle:
        raise RuntimeError(
            "❌ Variable d'environnement VISUV_SECRET_KEY manquante.\n"
            "Générez une clé : python -c \"from cryptography.fernet import Fernet; "
            "print(Fernet.generate_key().decode())\"\n"
            "Puis définissez VISUV_SECRET_KEY dans votre environnement."
        )
    return Fernet(cle.encode())


def chiffrer(valeur: str) -> str:
    """Chiffre une valeur en clair → stockage sécurisé en base."""
    return _get_fernet().encrypt(valeur.encode()).decode()


def dechiffrer(valeur_chiffree: str) -> str:
    """Déchiffre une valeur lue depuis la base → valeur en clair."""
    return _get_fernet().decrypt(valeur_chiffree.encode()).decode()


def _est_chiffre(valeur: str) -> bool:
    """
    Détecte si une valeur est déjà chiffrée par Fernet.
    Un token Fernet commence toujours par 'gAAAAA' (base64 d'un byte de version 0x80).
    Utile pendant la migration pour ne pas double-chiffrer.
    """
    return valeur.startswith("gAAAAA")


# =========================================================
# 👤 CRUD UTILISATEURS
# =========================================================
def enregistrer_utilisateur(nom, username, telephone, email, mot_de_passe, role="Utilisateur"):
    conn     = get_connection()
    cursor   = conn.cursor()
    mdp_hash = hash_mot_de_passe(mot_de_passe)
    try:
        cursor.execute(
            """INSERT INTO utilisateurs (nom, username, telephone, email, mot_de_passe, role)
               VALUES (%s, %s, %s, %s, %s, %s)""",
            (nom, username, telephone, email, mdp_hash, role)
        )
        conn.commit()
        return True, "✅ Inscription réussie !"
    except mysql.connector.Error as err:
        if err.errno == 1062:
            if "email"     in str(err): return False, "❌ Cet email est déjà utilisé."
            if "telephone" in str(err): return False, "❌ Ce numéro de téléphone est déjà utilisé."
            return False, "❌ Ce compte existe déjà."
        return False, f"❌ Erreur base de données : {err}"
    finally:
        cursor.close()
        conn.close()


def verifier_identifiants(username, password):
    conn     = get_connection()
    cursor   = conn.cursor(dictionary=True)
    mdp_hash = hashlib.sha256(password.encode()).hexdigest()
    try:
        cursor.execute(
            "SELECT role FROM utilisateurs WHERE BINARY username = %s AND BINARY mot_de_passe = %s",
            (username, mdp_hash)
        )
        utilisateur = cursor.fetchone()
        if utilisateur:
            return True, utilisateur["role"]
        return False, None
    except mysql.connector.Error as err:
        print(f"Erreur SQL : {err}")
        return False, None
    finally:
        cursor.close()
        conn.close()


def supprimer_utilisateur(user_id):
    conn   = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM utilisateurs WHERE id = %s", (user_id,))
        conn.commit()
    finally:
        cursor.close()
        conn.close()


def charger_utilisateurs():
    conn   = get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            """SELECT id, nom, username, telephone, email, role,
                      DATE_FORMAT(date_creation, '%d/%m/%Y') AS date_creation
               FROM utilisateurs"""
        )
        rows = cursor.fetchall()
        return pd.DataFrame(rows) if rows else pd.DataFrame(
            columns=["id", "nom", "username", "telephone", "email", "role", "date_creation"]
        )
    except Exception as e:
        st.error(f"❌ Erreur lors du chargement des utilisateurs : {e}")
        return pd.DataFrame(
            columns=["id", "nom", "username", "telephone", "email", "role", "date_creation"]
        )
    finally:
        cursor.close()
        conn.close()


def modifier_utilisateur(user_id, infos):
    conn   = get_connection()
    cursor = conn.cursor()
    try:
        if infos.get("mot_de_passe") and infos["mot_de_passe"].strip():
            mdp_hash = hash_mot_de_passe(infos["mot_de_passe"])
            cursor.execute(
                """UPDATE utilisateurs
                   SET nom=%s, telephone=%s, email=%s, role=%s, mot_de_passe=%s
                   WHERE id=%s""",
                (infos["nom"], infos["telephone"], infos["email"],
                 infos["role"], mdp_hash, user_id)
            )
        else:
            cursor.execute(
                """UPDATE utilisateurs
                   SET nom=%s, telephone=%s, email=%s, role=%s
                   WHERE id=%s""",
                (infos["nom"], infos["telephone"], infos["email"],
                 infos["role"], user_id)
            )
        conn.commit()
        return True
    except mysql.connector.Error as err:
        st.error(f"❌ Erreur MySQL : {err}")
        return False
    finally:
        cursor.close()
        conn.close()


def get_user_by_username(username):
    conn   = get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            """SELECT id, nom, username, telephone, email, role,
                      DATE_FORMAT(date_creation, '%d/%m/%Y à %H:%i') AS date_creation
               FROM utilisateurs WHERE username = %s""",
            (username,)
        )
        return cursor.fetchone()
    except Exception:
        return None
    finally:
        cursor.close()
        conn.close()


# =========================================================
# 🔧 UTILITAIRES SESSION
# =========================================================
def est_connecte():
    return (
        "utilisateur_connecte" in st.session_state
        and st.session_state["utilisateur_connecte"] is not None
    )


# =========================================================
# 🌐 GESTION DES COMPTES CGAWEB
# =========================================================

def charger_comptes():
    """
    Retourne tous les comptes CGAWEB pour l'affichage dans l'interface admin.
    Les champs password_cga et secret_otp sont DÉCHIFFRÉS pour que l'admin
    puisse les lire et les modifier.
    Si le déchiffrement échoue (valeur en clair legacy), la valeur brute est retournée.
    """
    conn   = get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM comptes_cgaweb")
        comptes = cursor.fetchall()
        for c in comptes:
            # Déchiffrement sécurisé — ne plante pas si la valeur est en clair (legacy)
            for champ in ("password_cga", "secret_otp"):
                try:
                    if _est_chiffre(c[champ]):
                        c[champ] = dechiffrer(c[champ])
                except Exception:
                    pass  # valeur en clair ou corrompue — on la laisse telle quelle
        return comptes
    finally:
        cursor.close()
        conn.close()


def enregistrer_nouveau_compte(user: str, password: str, secret: str) -> bool:
    """
    Enregistre un nouveau compte CGAWEB.
    password et secret sont chiffrés avant écriture en base.
    """
    conn   = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """INSERT INTO comptes_cgaweb (user_cga, password_cga, secret_otp)
               VALUES (%s, %s, %s)""",
            (user, chiffrer(password), chiffrer(secret))
        )
        conn.commit()
        return True
    except Exception as e:
        print(f"❌ Erreur enregistrement compte : {e}")
        return False
    finally:
        cursor.close()
        conn.close()


def modifier_compte(compte_id: int, user: str, password: str, secret: str) -> bool:
    """
    Met à jour un compte CGAWEB existant.
    password et secret sont toujours rechiffrés à l'écriture.
    """
    conn   = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """UPDATE comptes_cgaweb
               SET user_cga = %s, password_cga = %s, secret_otp = %s
               WHERE id = %s""",
            (user, chiffrer(password), chiffrer(secret), compte_id)
        )
        conn.commit()
        return True
    except Exception as e:
        print(f"❌ Erreur modification compte : {e}")
        return False
    finally:
        cursor.close()
        conn.close()


def marquer_compte_actif(compte_id: int) -> None:
    """Désactive tous les comptes puis active uniquement celui demandé."""
    conn   = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("UPDATE comptes_cgaweb SET est_actif = FALSE")
        cursor.execute("UPDATE comptes_cgaweb SET est_actif = TRUE WHERE id = %s", (compte_id,))
        conn.commit()
    finally:
        cursor.close()
        conn.close()


def obtenir_compte_actif_worker() -> dict | None:
    """
    Retourne le compte actif avec password et secret DÉCHIFFRÉS.
    À utiliser partout où l'on a besoin des credentials en clair
    (workers, tests de connexion, etc.).
    """
    conn   = get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM comptes_cgaweb WHERE est_actif = TRUE LIMIT 1")
        compte = cursor.fetchone()
        if compte:
            try:
                if _est_chiffre(compte["password_cga"]):
                    compte["password_cga"] = dechiffrer(compte["password_cga"])
            except Exception as e:
                print(f"⚠️ Déchiffrement password_cga échoué : {e}")
            try:
                if _est_chiffre(compte["secret_otp"]):
                    compte["secret_otp"] = dechiffrer(compte["secret_otp"])
            except Exception as e:
                print(f"⚠️ Déchiffrement secret_otp échoué : {e}")
        return compte
    finally:
        cursor.close()
        conn.close()


def desactiver_tous_les_comptes() -> bool:
    """Désactive tous les comptes CGAWEB (aucun compte actif)."""
    conn   = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("UPDATE comptes_cgaweb SET est_actif = FALSE")
        conn.commit()
        return True
    except Exception as e:
        print(f"❌ Erreur désactivation SQL : {e}")
        return False
    finally:
        cursor.close()
        conn.close()


def supprimer_compte_sql(compte_id: int) -> None:
    conn   = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM comptes_cgaweb WHERE id = %s", (compte_id,))
        conn.commit()
    finally:
        cursor.close()
        conn.close()


# =========================================================
# 🔄 MIGRATION — à exécuter UNE SEULE FOIS sur les données legacy
# =========================================================
def migrer_comptes_vers_chiffrement() -> None:
    """
    Chiffre les comptes existants dont les valeurs sont encore en clair.
    Idempotent : détecte les valeurs déjà chiffrées et les ignore.
    Supprimer cette fonction après migration confirmée.
    """
    conn   = get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT id, password_cga, secret_otp FROM comptes_cgaweb")
        comptes = cursor.fetchall()
        for c in comptes:
            pwd_final    = c["password_cga"] if _est_chiffre(c["password_cga"]) else chiffrer(c["password_cga"])
            secret_final = c["secret_otp"]   if _est_chiffre(c["secret_otp"])   else chiffrer(c["secret_otp"])
            cursor.execute(
                "UPDATE comptes_cgaweb SET password_cga = %s, secret_otp = %s WHERE id = %s",
                (pwd_final, secret_final, c["id"])
            )
            statut_pwd    = "déjà chiffré" if _est_chiffre(c["password_cga"]) else "chiffré"
            statut_secret = "déjà chiffré" if _est_chiffre(c["secret_otp"])   else "chiffré"
            print(f"  Compte {c['id']} — password: {statut_pwd} | secret: {statut_secret}")
        conn.commit()
        print("✅ Migration terminée.")
    except Exception as e:
        print(f"❌ Erreur migration : {e}")
    finally:
        cursor.close()
        conn.close()
#-------------------Fonction pour réactivation et réinitialisation du code parental-------------------

def action_abonne(driver, fiche_abonne, action="reactivation", timeout=10):
    """
    Exécute une action sur la fiche abonné CGAWEB.

    Paramètres :
        driver : Selenium WebDriver connecté
        fiche_abonne : WebElement de la fiche abonné ouverte
        action : "reactivation" ou "reinit"
        timeout : temps d'attente pour le formulaire (par défaut 10s)

    Retour :
        (True, None) si succès
        (False, message) si échec (alerte ou erreur)
    """
    # --- mapping des boutons ---
    boutons = {
        "reactivation": {"name": "reactivation", "value": "Réactivation"},
        "reinit": {"name": "reinit", "value": "Reinit. Code Parental"}
    }

    if action not in boutons:
        return (False, f"Action '{action}' non reconnue")

    try:
        # 1. Cliquer sur le lien "Activité" dans la fiche abonné
        try:
            bouton_activite = fiche_abonne.find_element(
                By.XPATH, "//a[@title='Activité' and contains(@href, 'modActivity.do')]"
            )
            bouton_activite.click()
        except NoSuchElementException:
            return (False, "Bouton 'Activité' introuvable dans la fiche abonné")

        # 2. Attendre que le formulaire modactivity soit visible
        try:
            WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((By.NAME, "modactivity"))
            )
        except TimeoutException:
            return (False, "Le formulaire modactivity n'est pas apparu")

        # 3. Cliquer sur le bouton correspondant à l'action
        try:
            bouton = driver.find_element(
                By.XPATH,
                f"//input[@name='{boutons[action]['name']}' and @value=\"{boutons[action]['value']}\"]"
            )
            bouton.click()
            time.sleep(2)  # attendre la réaction du site
        except NoSuchElementException:
            return (False, f"Bouton '{boutons[action]['value']}' introuvable")

        # 4. Vérifier si une alerte apparaît
        try:
            WebDriverWait(driver, 3).until(EC.alert_is_present())
            alert = driver.switch_to.alert
            message = alert.text
            alert.accept()  # fermer l'alerte
            return (False, message)  # échec avec message
        except TimeoutException:
            return (True, None)  # pas d'alerte → succès

    except Exception as e:
        return (False, f"Erreur pendant l'action '{action}' : {e}")

#----------------------------------------
#fonction pour le traitement des prestattion de service sur cgaweb
#----------------------------------------

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time

from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

def extraire_prestations_service(driver, timeout=15):
    prestations = []

    try:
        # ================== MENU ==================
        driver.switch_to.default_content()
        WebDriverWait(driver, timeout).until(
            EC.frame_to_be_available_and_switch_to_it("_left")
        )
        print("✅ Frame _left chargée")

        menu_abonne = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable(
                (By.XPATH, '//a[contains(@onmouseover,"ABOM0000SHOP")]')
            )
        )
        menu_abonne.click()

        driver.execute_script("showMenuItem(window, 'ABOM0300SHOP', 0);")
        time.sleep(0.5)

        prestations_link = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.ID, "MODSUBSCRIBERSERVICE"))
        )
        driver.execute_script("arguments[0].click();", prestations_link)
        print("✅ Accès Prestations-Services")

        # ================== PAGE ==================
        driver.switch_to.default_content()
        WebDriverWait(driver, timeout).until(
            EC.frame_to_be_available_and_switch_to_it("_right")
        )
        print("✅ Frame _right chargée")

        # ================== AUCUNE PRESTATION ==================
        if "Pas de prestation enregistrée pour cet abonné" in driver.page_source:
            print("ℹ️ Aucune prestation enregistrée")
            return []

        # ================== TABLE ==================
        tables = driver.find_elements(By.TAG_NAME, "table")
        table_prestations = None

        for table in tables:
            rows = table.find_elements(By.TAG_NAME, "tr")
            if len(rows) > 1:
                table_prestations = table
                break

        if not table_prestations:
            print("ℹ️ Aucune table de prestations détectée")
            return []

        rows = table_prestations.find_elements(By.TAG_NAME, "tr")
        print(f"📊 {len(rows)-1} lignes de prestations")

        for tr in rows[1:]:
            tds = tr.find_elements(By.TAG_NAME, "td")
            if len(tds) < 6:
                continue

            prestations.append({
                "Prestation": tds[0].text.strip(),
                "Libellé": tds[1].text.strip(),
                "Distributeur": tds[2].text.strip(),
                "Antenniste": tds[3].text.strip(),
                "Validation": tds[4].text.strip().lower() == "oui",
                "Annulation": tds[5].text.strip().lower() == "oui",
                "Statut": _calcul_statut(tds)
            })

        print(f"✅ {len(prestations)} prestations extraites")
        return prestations

    except Exception as e:
        print(f"❌ Erreur extraction prestations : {type(e).__name__}: {e}")

        # ================== DEBUG ==================
        try:
            driver.save_screenshot("debug_erreur_prestations.png")
            with open("debug_prestations.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            print("📸 Screenshot + HTML sauvegardés")
        except:
            pass

        # ================== DÉCONNEXION ==================
        deconnexion_cgaweb(driver)
        return []


def _calcul_statut(tds):
    prestation = tds[0].text.strip().upper()
    validation = tds[4].text.strip().lower() == "oui"
    annulation = tds[5].text.strip().lower() == "oui"
    antenniste = tds[3].text.strip()

    if prestation == "SPSC":
        return "SPCSC"
    if annulation:
        return "ANNULÉ"
    if validation:
        return "TOK"
    if antenniste:
        return "ACCEPTER"
    return "A PLANIFIER"


import time
import re
import io
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    WebDriverException,
    StaleElementReferenceException,
)

# ============================================================
# GESTION ARRÊT
# ============================================================
STOP_REQUESTED = False

def demander_arret():
    global STOP_REQUESTED
    STOP_REQUESTED = True

def reset_arret():
    global STOP_REQUESTED
    STOP_REQUESTED = False

def arret_demande():
    return STOP_REQUESTED


# ============================================================
# CONNEXION
# ============================================================
def connexion_savant(driver_instance, url, username, password, timeout=30):
    """Connexion à l'application SAVANT — version robuste Streamlit Cloud."""
    wait = WebDriverWait(driver_instance, timeout)
    
    try:
        # 1. Chargement de la page avec retry
        print(f"🌐 Chargement de : {url}")
        driver_instance.get(url)

        # 2. Attendre que la page soit vraiment chargée (pas juste DOM ready)
        wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
        print(f"📄 Page chargée : {driver_instance.title}")

        # 3. Champ identifiant
        champ_identifiant = wait.until(
            EC.presence_of_element_located((By.ID, "in_username"))
        )
        wait.until(EC.element_to_be_clickable((By.ID, "in_username")))
        champ_identifiant.clear()
        champ_identifiant.send_keys(username)
        print("✏️ Identifiant saisi")

        # 4. Champ mot de passe
        champ_mot_de_passe = wait.until(
            EC.presence_of_element_located((By.ID, "in_password"))
        )
        wait.until(EC.element_to_be_clickable((By.ID, "in_password")))
        champ_mot_de_passe.clear()
        champ_mot_de_passe.send_keys(password)
        print("🔑 Mot de passe saisi")

        # 5. Bouton connexion — double stratégie de localisation
        try:
            bouton_connexion = wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//input[@type='submit' and contains(@class, 'newimgbtn')]")
                )
            )
        except TimeoutException:
            # Fallback si le sélecteur principal échoue
            print("⚠️ Sélecteur principal échoué, tentative fallback...")
            bouton_connexion = wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//input[@type='submit']")
                )
            )

        # 6. Clic via JavaScript (plus fiable en headless)
        driver_instance.execute_script("arguments[0].click();", bouton_connexion)
        print("🖱️ Bouton cliqué")

        # 7. Attente confirmation connexion
        wait.until(EC.presence_of_element_located((By.ID, "menudiv")))
        print("✅ Connexion réussie.")
        return True, driver_instance, wait

    except TimeoutException as e:
        print(f"❌ Timeout — page courante : {driver_instance.current_url}")
        print(f"   Titre : {driver_instance.title}")
        # Screenshot pour debug (Streamlit Cloud écrit dans /tmp)
        try:
            screenshot_path = "/tmp/debug_connexion.png"
            driver_instance.save_screenshot(screenshot_path)
            print(f"📸 Screenshot sauvegardé : {screenshot_path}")
        except Exception:
            pass
        return False, None, None

    except (NoSuchElementException, WebDriverException) as e:
        print(f"❌ Erreur connexion ({e.__class__.__name__}): {e}")
        return False, None, None

# ============================================================
# NAVIGATION
# ============================================================
def naviguer_page_intervention(driver_instance, wait):
    """Clique sur le lien 'Intervention' pour naviguer vers la liste."""
    try:
        driver_instance.switch_to.default_content()
        wait.until(EC.presence_of_element_located((By.ID, "menudiv")))
        intervention_btn = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//a[contains(@href, 'INTER_PENDING')]"))
        )
        driver_instance.execute_script("arguments[0].click();", intervention_btn)
        wait.until(EC.presence_of_element_located((By.ID, "home_body")))
        print("✅ Page Intervention chargée.")
        return True
    except (TimeoutException, StaleElementReferenceException) as e:
        print(f"❌ Erreur navigation : {e.__class__.__name__}")
        return False


# ============================================================
# SÉLECTION DES DATES
# ============================================================
def selectionner_date_avec_validation(driver, wait, champ_id, date_str, valider=True):
    """Sélectionne une date dans le datepicker. Si valider=False, ne clique pas sur Valider."""
    try:
        jour, mois, annee = map(int, date_str.split("/"))
        mois_index = mois - 1
    except ValueError:
        print(f"❌ Format de date incorrect : {date_str}. Utiliser JJ/MM/AAAA.")
        return False

    MOIS_TEXT = [
        "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre",
    ]

    # 1. Ouvrir le datepicker
    try:
        champ = wait.until(EC.element_to_be_clickable((By.ID, champ_id)))
        driver.execute_script("arguments[0].click();", champ)
        time.sleep(0.5)
    except TimeoutException:
        print(f"❌ Champ {champ_id} non cliquable")
        return False

    # 2. Attendre visibilité
    try:
        wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "ui-datepicker-title")))
    except TimeoutException:
        print("❌ Datepicker non visible")
        return False

    # 3. Naviguer vers le bon mois/année
    max_iterations = 24
    for _ in range(max_iterations):
        try:
            nav_prev = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "ui-icon-circle-triangle-w")))
            nav_next = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "ui-icon-circle-triangle-e")))
            titre = driver.find_element(By.CLASS_NAME, "ui-datepicker-title")
            mois_affiche = titre.find_element(By.CLASS_NAME, "ui-datepicker-month").text
            annee_affiche = int(titre.find_element(By.CLASS_NAME, "ui-datepicker-year").text)
            mois_actuel = MOIS_TEXT.index(mois_affiche)
        except (NoSuchElementException, StaleElementReferenceException, ValueError) as e:
            print(f"❌ Erreur navigation calendrier : {e}")
            return False

        if annee_affiche < annee or (annee_affiche == annee and mois_actuel < mois_index):
            nav_next.click()
        elif annee_affiche > annee or (annee_affiche == annee and mois_actuel > mois_index):
            nav_prev.click()
        else:
            break
        time.sleep(0.2)
    else:
        print("❌ Limite de navigation atteinte")
        return False

    # 4. Cliquer sur le jour
    try:
        xpath_jour = f"//table[@class='ui-datepicker-calendar']//a[text()='{jour}']"
        jour_element = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_jour)))
        jour_element.click()
        time.sleep(0.3)
        valeur = driver.find_element(By.ID, champ_id).get_attribute("value")
        print(f"  ✓ Date sélectionnée : {valeur}")
    except TimeoutException:
        print(f"❌ Jour {jour} introuvable")
        return False

    # 5. Valider si demandé
    if valider:
        try:
            bouton_valider = wait.until(EC.element_to_be_clickable((By.ID, "btn_period_valid")))
            driver.execute_script("arguments[0].click();", bouton_valider)
            time.sleep(1)
            print(f"✅ Date {date_str} validée")
        except TimeoutException:
            print("❌ Bouton Valider non cliquable")
            return False

    return True


def activer_et_selectionner_dates(driver, wait, date_debut, date_fin,
                                   champ_debut_id="intervention_from_datecrea",
                                   champ_fin_id="intervention_to_datecrea",
                                   div_dates_id="intervention_date_div",
                                   bouton_valider_id="btn_period_valid"):
    """Active le sélecteur de dates et configure la période complète."""
    # Ouvrir calendrier début
    try:
        champ_from = wait.until(EC.element_to_be_clickable((By.ID, champ_debut_id)))
        champ_from.click()
        wait.until(EC.visibility_of_element_located((By.ID, div_dates_id)))
        print("✅ Calendrier visible pour date début")
    except TimeoutException:
        print("❌ Impossible d'activer le champ date début")
        return False

    if not selectionner_date_avec_validation(driver, wait, champ_debut_id, date_debut, valider=False):
        print("❌ Sélection date début échouée")
        return False

    # Ouvrir calendrier fin
    try:
        champ_to = wait.until(EC.element_to_be_clickable((By.ID, champ_fin_id)))
        champ_to.click()
        time.sleep(0.3)
        print("✅ Calendrier visible pour date fin")
    except TimeoutException:
        print("❌ Impossible d'activer le champ date fin")
        return False

    if not selectionner_date_avec_validation(driver, wait, champ_fin_id, date_fin, valider=False):
        print("❌ Sélection date fin échouée")
        return False

    # Valider les deux dates ensemble
    try:
        bouton_valider = wait.until(EC.element_to_be_clickable((By.ID, bouton_valider_id)))
        driver.execute_script("arguments[0].click();", bouton_valider)
        time.sleep(1)
        print(f"✅ Période validée : {date_debut} → {date_fin}")
        return True
    except TimeoutException:
        print("❌ Bouton Valider non cliquable")
        return False


# ============================================================
# SÉLECTION DES STATUTS
# ============================================================
def selectionner_statuts(driver, wait, statuts_choisis):
    """Sélectionne les statuts dans le dropdown 'Intervention Status'."""
    DROPDOWN_ID = "intervention_status_select-button"
    MENU_ID = "intervention_status_select-menu"
    VALIDER_ID = "btn_period_valid"

    def cliquer_option(texte_option):
        for _ in range(3):
            try:
                xpath_option = f"//ul[@id='{MENU_ID}']/li[contains(normalize-space(.), '{texte_option}')]"
                option = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_option)))
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", option)
                option.click()
                return True
            except StaleElementReferenceException:
                time.sleep(0.2)
            except TimeoutException:
                print(f"⚠ Option introuvable : {texte_option}")
                return False
        return False

    try:
        dropdown = wait.until(EC.element_to_be_clickable((By.ID, DROPDOWN_ID)))
        driver.execute_script("arguments[0].click();", dropdown)
        time.sleep(0.3)
    except TimeoutException:
        print(f"❌ Impossible d'ouvrir le dropdown : {DROPDOWN_ID}")
        return

    for statut in statuts_choisis:
        cliquer_option(statut)
        time.sleep(0.2)

    try:
        bouton_valider = wait.until(EC.element_to_be_clickable((By.ID, VALIDER_ID)))
        driver.execute_script("arguments[0].click();", bouton_valider)
        time.sleep(1)
        print("✅ Statuts sélectionnés et filtre validé.")
    except TimeoutException:
        print(f"❌ Bouton Valider non cliquable : {VALIDER_ID}")


def selectionner_statut_temporaire(driver, wait):
    """Sélectionne uniquement le statut Temporaire."""
    try:
        dropdown = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "#intervention_status_select-button"))
        )
        dropdown.click()
        option_temp = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//li[contains(., 'Temporaire')]"))
        )
        option_temp.click()
        wait.until(EC.presence_of_element_located((By.ID, "tbl_inter_pending")))
        return True
    except Exception as e:
        print("Erreur sélection statut Temporaire :", e)
        return False


# ============================================================
# EXTRACTION DU TABLEAU (autres statuts)
# ============================================================
def extraire_tableau(driver, wait, statuts_choisis):
    """Extraction complète et fiable de toutes les pages."""
    resultats = []
    numeros_vus = set()

    # ── Aller à la première page ──────────────────────────
    try:
        wait.until(EC.element_to_be_clickable((By.ID, "tbl_inter_pending_first"))).click()
        time.sleep(2)
    except Exception:
        pass

    page_count = 0

    while True:
        if arret_demande():
            print("⛔ Arrêt demandé")
            break

        page_count += 1

        # ── Attendre que le tableau soit prêt ─────────────
        try:
            wait.until(EC.invisibility_of_element_located(
                (By.ID, "tbl_inter_pending_processing")
            ))
        except Exception:
            pass

        try:
            wait.until(EC.presence_of_element_located(
                (By.XPATH, "//table[@id='tbl_inter_pending']/tbody/tr")
            ))
        except Exception:
            pass

        time.sleep(0.5)

        # ── Extraire TOUTES les lignes de la page ─────────
        lignes = driver.find_elements(
            By.XPATH,
            "//table[@id='tbl_inter_pending']/tbody/tr"
            "[not(contains(@class,'dataTables_empty'))]"
        )

        nb_lignes = len(lignes)
        print(f"📄 Page {page_count} — {nb_lignes} lignes")

        if nb_lignes == 0:
            print("  → Tableau vide, arrêt")
            break

        # ── Lire chaque ligne avec retry anti-stale ───────
        for idx in range(nb_lignes):
            if arret_demande():
                break

            for tentative in range(3):
                try:
                    lignes_actuelles = driver.find_elements(
                        By.XPATH,
                        "//table[@id='tbl_inter_pending']/tbody/tr"
                        "[not(contains(@class,'dataTables_empty'))]"
                    )

                    if idx >= len(lignes_actuelles):
                        break

                    tr = lignes_actuelles[idx]
                    colonnes = tr.find_elements(By.TAG_NAME, "td")

                    if len(colonnes) < 5:
                        break

                    numero = colonnes[4].text.strip()

                    if not numero:
                        break

                    if numero not in numeros_vus:
                        numeros_vus.add(numero)
                        resultats.append({
                            "Numero": numero,
                            "Statut": ", ".join(statuts_choisis)
                        })

                    break

                except StaleElementReferenceException:
                    print(f"  ⚠️ Stale ligne {idx}, tentative {tentative + 1}/3")
                    time.sleep(0.3)
                    continue
                except Exception as e:
                    print(f"  ⚠️ Erreur ligne {idx} : {e}")
                    break

        print(f"  → Cumulé : {len(resultats)} interventions extraites")

        # ── Pagination sans référence à un WebElement ─────
        try:
            # ✅ Capturer le texte AVANT le clic (string pure, pas WebElement)
            lignes_avant = driver.find_elements(
                By.XPATH,
                "//table[@id='tbl_inter_pending']/tbody/tr"
                "[not(contains(@class,'dataTables_empty'))]"
            )
            # ✅ Extraire le texte immédiatement en string pure
            texte_avant = lignes_avant[0].find_elements(
                By.TAG_NAME, "td"
            )[4].text.strip() if lignes_avant else ""

            # Vérifier si Next est disponible
            next_btn = driver.find_element(By.ID, "tbl_inter_pending_next")
            classes = next_btn.get_attribute("class") or ""

            if "disabled" in classes or "ui-state-disabled" in classes:
                print("✅ Dernière page atteinte")
                break

            # Clic JS
            driver.execute_script("arguments[0].click();", next_btn)

            # ✅ Attendre en comparant uniquement des strings (pas de WebElement)
            for _ in range(30):  # max 15 secondes
                time.sleep(0.5)
                try:
                    nouvelles_lignes = driver.find_elements(
                        By.XPATH,
                        "//table[@id='tbl_inter_pending']/tbody/tr"
                        "[not(contains(@class,'dataTables_empty'))]"
                    )
                    if not nouvelles_lignes:
                        continue
                    texte_apres = nouvelles_lignes[0].find_elements(
                        By.TAG_NAME, "td"
                    )[4].text.strip()

                    if texte_apres != texte_avant:
                        break  # ✅ page changée
                except Exception:
                    continue

        except Exception as e:
            print(f"  → Fin pagination ({e})")
            break

    print(f"\n🏁 TOTAL : {len(resultats)} interventions — Statuts : {statuts_choisis}")
    return resultats


# ============================================================
# EXTRACTION DES TEMPORAIRES
# ============================================================
def extraire_interventions_temporaire(driver, wait):
    """Extrait les interventions de statut Temporaire (ouvre chaque fiche)."""
    resultats = []
    decodeurs_vus = set()
    page_num = 1

    while True:
        if arret_demande():
            print("⛔ Arrêt demandé – retour résultats temporaires")
            break

        wait.until(EC.presence_of_element_located((By.ID, "tbl_inter_pending")))
        voir_btns = driver.find_elements(By.XPATH, "//a[contains(text(),'Voir')]")

        for i, btn in enumerate(voir_btns, start=1):
            if arret_demande():
                break
            try:
                driver.execute_script("arguments[0].click();", btn)
                wait.until(lambda d: len(d.window_handles) > 1)
                driver.switch_to.window(driver.window_handles[-1])
                wait.until(EC.presence_of_all_elements_located(
                    (By.CSS_SELECTOR, "h3.ui-accordion-header"))
                )

                headers = driver.find_elements(By.CSS_SELECTOR, "h3.ui-accordion-header")
                header_rdv = header_cr = None
                for h in headers:
                    t = h.text.lower()
                    if "rdv" in t:
                        header_rdv = h
                    if "compte" in t or "rendu" in t:
                        header_cr = h

                id_tech = None
                if header_rdv:
                    header_rdv.click()
                    time.sleep(0.5)
                    try:
                        el = driver.find_element(
                            By.XPATH, "//*[contains(text(),'ID Tech')]/following::div[1]"
                        )
                        id_tech = el.text.strip()
                    except Exception:
                        pass
                    header_rdv.click()

                decodeurs = []
                if header_cr:
                    header_cr.click()
                    time.sleep(0.5)
                    inputs = driver.find_elements(
                        By.XPATH, "//input[starts-with(@id,'ref_decodeur')]"
                    )
                    for inp in inputs:
                        val = inp.get_attribute("value").strip()
                        if re.fullmatch(r"\d{14}", val) and val not in decodeurs_vus:
                            decodeurs.append(val)
                            decodeurs_vus.add(val)
                    header_cr.click()

                if decodeurs:
                    resultats.append({
                        "Page": page_num,
                        "Ligne": i,
                        "ID Tech": id_tech,
                        "Décodeurs": ", ".join(decodeurs),
                    })

                driver.close()
                driver.switch_to.window(driver.window_handles[0])

            except Exception as e:
                print("⚠️ Erreur ligne temporaire :", e)

        try:
            next_btn = driver.find_element(By.XPATH, "//a[contains(text(),'Suivant')]")
            if "ui-state-disabled" in next_btn.get_attribute("class"):
                break
            next_btn.click()
            page_num += 1
            time.sleep(1.5)
        except Exception:
            break

    return resultats


# ============================================================
# EXPORT EXCEL MULTI-FEUILLES
# ============================================================
def generer_excel_multi_feuilles(data_temporaire, data_autres):
    """Génère un fichier Excel avec deux feuilles : TEMPORAIRE et AUTRES STATUTS."""
    wb = Workbook()

    # Feuille TEMPORAIRE
    ws_t = wb.active
    ws_t.title = "TEMPORAIRE"
    if data_temporaire:
        ws_t.append(list(data_temporaire[0].keys()))
        for row in data_temporaire:
            ws_t.append(list(row.values()))
    else:
        ws_t.append(["Aucune donnée TEMPORAIRE"])

    # Feuille AUTRES STATUTS
    ws_a = wb.create_sheet(title="AUTRES STATUTS")
    if data_autres:
        ws_a.append(list(data_autres[0].keys()))
        for row in data_autres:
            ws_a.append(list(row.values()))
    else:
        ws_a.append(["Aucune donnée AUTRES STATUTS"])

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


# ============================================================
# UTILITAIRES
# ============================================================
def est_statut_temporaire(statuts):
    """Vérifie si 'Temporaire' est dans la liste des statuts sélectionnés."""
    return "Temporaire" in statuts


def connexion_savant(driver_instance, url, username, password, timeout=30):
    """Connexion à l'application SAVANT"""
    wait = WebDriverWait(driver_instance, timeout)
    try:
        driver_instance.get(url)
        # Attente des champs de connexion
        champ_identifiant = wait.until(EC.element_to_be_clickable((By.ID, "in_username")))
        champ_mot_de_passe = wait.until(EC.element_to_be_clickable((By.ID, "in_password")))
        
        champ_identifiant.send_keys(username)
        champ_mot_de_passe.send_keys(password)
        
        # Attente et clic sur le bouton de connexion
        bouton_connexion = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//input[@type='submit' and contains(@class, 'newimgbtn')]"))
        )
        bouton_connexion.click()
        
        # ✅ Attente d'un élément post-connexion fiable (ex: la barre de menu)
        wait.until(EC.presence_of_element_located((By.ID, "menudiv"))) 
        print("✅ Connexion réussie.")
        return True, driver_instance, wait
        
    except (TimeoutException, NoSuchElementException, WebDriverException) as e:
        print(f"❌ Erreur connexion ({e.__class__.__name__}): {e}")
        return False, None, None
